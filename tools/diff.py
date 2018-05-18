import difflib
import os
# 前提として、同名ファイルがある場合は、正常に動作しない
import re
from operator import attrgetter

import openpyxl as px


def diff(target_path_1, target_path_2):
  target_files_1 = __flat_files(target_path_1)
  target_files_2 = __flat_files(target_path_2)

  with open('diff.txt', 'w') as fw:
    wb = px.Workbook()
    row = 2
    ws = wb.create_sheet(title='result', index=0)  # シート名を設定して、左から2番目の位置に追加
    ws[f'A{str(row)}'] = 'target1'
    ws[f'B{str(row)}'] = 'line'
    ws[f'C{str(row)}'] = 'content'
    ws[f'D{str(row)}'] = 'type'
    ws[f'E{str(row)}'] = 'target2'
    ws[f'F{str(row)}'] = 'line'
    ws[f'G{str(row)}'] = 'content'
    for name, path in target_files_1.items():
      if name not in target_files_2:
        continue

      block_starting = False
      target1_starting = False
      block_diff_results = []
      target1_start_row = 0
      target2_start_row = 0

      try:
        d = difflib.HtmlDiff()
        result = d.make_table(__read_lines(path), __read_lines(target_files_2[name]), context=True)
        if 'No Differences Found' not in result:
          with open("./output/" + name + ".html", "w", encoding='utf-8') as fw_html:
            fw_html.writelines(result)
      except UnicodeDecodeError:
        continue

      for line in __context_diff(path, target_files_2[name]):
        fw.write(line)
        if re.match('^[\*]{15}', line):
          if block_starting:
            sorted_block_diff_results = sorted(block_diff_results, key=attrgetter('row', 'target'))
            for result in sorted_block_diff_results:
              if result.target == 1:
                row += 1
                ws[f'A{str(row)}'] = path
                ws[f'B{str(row)}'] = result.row
                ws[f'C{str(row)}'] = result.line[1:]
                ws[f'D{str(row)}'] = __convert_to_type(result.line[0])
              else:
                if '!' not in result.line[0]:
                  row += 1
                  ws[f'D{str(row)}'] = __convert_to_type(result.line[0])
                ws[f'E{str(row)}'] = target_files_2[name]
                ws[f'F{str(row)}'] = result.row
                ws[f'G{str(row)}'] = result.line[1:]
            block_diff_results = []

          block_starting = False if block_starting else True
          continue

        target1_re = re.compile('^[\*]{3} [\d]+')
        target2_re = re.compile('^[-]{3} [\d]+')
        if target1_re.match(line):
          target1_starting = True
          rows = line.split(',')
          target1_start_row = int(rows[0].replace(' ', '').replace('*', ''))
          continue
        if target2_re.match(line):
          target1_starting = False
          rows = line.split(',')
          target2_start_row = int(rows[0].replace(' ', '').replace('-', ''))
          continue
        if re.match('^([\*]{3}|[-]{3})', line): continue

        if 0 < len(line) and re.match('^[!\+-]', line[0]):
          block_diff_results.append(
            BlockDiffResult(1, target1_start_row, line) if target1_starting else BlockDiffResult(2, target2_start_row,
                                                                                                 line))
        if target1_starting:
          target1_start_row += 1
        else:
          target2_start_row += 1

    wb.save('diff.xlsx')
    wb.close()


def __flat_files(root_dir):
  target_files_1 = {}
  for path, file in __find(root_dir):
    if os.path.isdir(os.path.join(path, file)):
      continue
    target_files_1[file] = os.path.join(path, file)
  return target_files_1


def __find(root_dir):
  for root, dirs, files in os.walk(root_dir):
    yield '', root
    for file in files:
      yield root, file


def __read_lines(path):
  with open(path, 'r', encoding='utf-8') as fr:
    return fr.readlines()


def __context_diff(path1, path2):
  try:
    return difflib.context_diff(
      __read_lines(path1),
      __read_lines(path2),
      fromfile=path1,
      tofile=path2)
  except UnicodeDecodeError:
    return ""


def __convert_to_type(line):
  return '!=' if '!' in line else '+'


class BlockDiffResult:
  def __init__(self, target, row, line):
    self.target = target
    self.row = row
    self.line = line

  def __repr__(self):
    return repr((self.target, self.row, self.line))
